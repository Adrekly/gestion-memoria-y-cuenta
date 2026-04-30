import openpyxl
from io import BytesIO
from typing import Optional
from datetime import datetime

async def generar_excel_inventario(db, sede: Optional[str] = None) -> BytesIO:
    """Genera el inventario de bienes en formato Excel (.xlsx)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario de Bienes"

    # Estilos básicos
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="003366", end_color="003366", fill_type="solid")

    # Título y metadatos
    ws.append(["UNIVERSIDAD NACIONAL EXPERIMENTAL DE GUAYANA"])
    ws.append(["Departamento de Bienes Publicos - Inventario General"])
    ws.append([f"Fecha de reporte: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
    if sede:
        ws.append([f"Filtro aplicado: Sede {sede}"])
    ws.append([])

    # Encabezados
    headers = [
        "Codigo Inventario", "Codigo SUDEBIP", "Descripcion", "Marca", "Modelo",
        "Serial", "Estado", "Condicion", "Sede", "Ubicacion Especifica",
        "Responsable", "Valor Adquisicion ($)"
    ]
    ws.append(headers)

    # Aplicar estilo al header
    header_row = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = header_font
        cell.fill = header_fill

    # Filtro
    filtro = {}
    if sede:
        filtro["sede.codigo"] = sede

    # Datos
    async for bien in db.bienes.find(filtro).sort("codigo_inventario", 1):
        ws.append([
            bien.get("codigo_inventario", ""),
            bien.get("codigo_sudebip", ""),
            bien.get("descripcion", ""),
            bien.get("marca", ""),
            bien.get("modelo", ""),
            bien.get("serial", ""),
            bien.get("estado", ""),
            bien.get("condicion", ""),
            bien.get("sede", {}).get("nombre", ""),
            bien.get("ubicacion_especifica", ""),
            bien.get("responsable", ""),
            bien.get("valor_adquisicion", 0)
        ])

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        if adjusted_width > 50:
            adjusted_width = 50
        ws.column_dimensions[column].width = adjusted_width

    # Guardar a buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
